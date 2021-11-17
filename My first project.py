command = ""
started = False
while True:
    command = input("> ").lower()
    if command == "start":
        if started:
            print('the car alreay started!')
        else:
            started = True
            print('car started')
    elif command  == 'stop':
        if not started:
            print('car already stopped!')
        else:
            started = False
            print('car stoped')
    elif command == "help":
        print("""
        start to go
        stop
        quit
        """)
    elif command == 'quit':
        break
    else:
        print('sorry i cant understand!')


prices = [10,20,30]

total = 0
for price in prices:
    total += price
print(f"total: {total}")


for x in range(4):
    for y in range(3):
        print((x,y))


numbers = [5, 2, 5, 2, 2]
for item in numbers:
    print('x'* item)


numbers = ['1', '2', '3', '4']
max = numbers[0]
for number in numbers:
    if number > max:
        max = number
print(max)


matrix = [
    [1, 2, 3],
    [4, 5, 6],
    [7, 8, 9]
]
for row in matrix:
    for item in row:
        print(item)


matrix = [
    [1,2,3],
    [4,5,6],
    [7,8,9],


]
 for row in matrix:
     for item in row:
         print(item)



numbers = [2, 4, 6, 8, 10]
numbers.append(30)
print(numbers)

numbers = [2, 4, 6, 8, 10]
numbers.insert(0,20)
print(numbers)

numbers = [2, 4, 6, 8, 10]
numbers.remove(10)
print(numbers)

numbers = [2, 4, 6, 8, 10]
numbers.clear()
print(numbers)

numbers = [2, 4, 6, 8, 10]
numbers.pop()
print(numbers)


numbers = [2, 4, 6, 8, 10]
print(numbers.index(10))

numbers = [2, 4, 6, 8, 10]
print(50 in numbers)

numbers = [2, 4, 6, 8, 6,  10]
numbers.sort()
numbers.reverse()
print(numbers)

numbers = [2, 4, 6, 8, 6,  10]
numbers2 = numbers.copy()
print(numbers2)

numbers = [2, 4, 6, 6, 8, 6, 10]
numbers2 = []
for number in numbers:
    if number not in numbers2:
        numbers2.append(number)
print(numbers2)


coordinates = (1, 2, 3)
x,y,z = coordinates
print(x)

coordinates = [1, 2, 3]
x,y,z = coordinates
print(x)

customer = {
    "name": "john smith",
    "age": 20,
    "is_veritify": True
}
customer["name"] = "jack smith"
print(customer["name"])



phone = input("phone: ")
digital_mapping = {
    "1": "one",
    "2": "two",
    "3": "three",
    "4": "four",
}
output = ""
for ch in phone:
    output += digital_mapping.get(ch, "!") + ""
print(output)

message = input("> ")
words = message.split(' ')
emojis = {
    "sunshine": "笑脸"
}
output = ""
for word in words:
    output += emojis.get(word, word) + " "
print(output)

def greet_user(first_name,last_name):
    print(f'hi {first_name} {last_name}!')
    print('welcome aboard')


print("start")
greet_user("john","smith")
print("finish")

def square(number):
    print(number * number)



print(square(3))


def emojis_converter(message):
    words = message.split(' ')
    emojis = {
        "sunshine": "笑脸"
    }
    output = ""
    for word in words:
        output += emojis.get(word, word) + " "
    return output



message = input(">")
print(emojis_converter(message))


try:
    age = float(input('age: '))
    print(age)
except ValueError:
    print('Invalid value')


class Point:
    def __init__(self,x,y):
        self.x = x
        self.y = y
    def move(self):
        print("move")

    def draw(self):
        print("draw")


point = Point(10,20)
print(point.x)


class Person:
    def __init__(self, name):
        self.name = name


     def talk(self):
        print(f"hi,i am {self.name}")



john = Person("john smith")
john.talk()

bob = Person("bob smith")
bob.talk()


class Mammal:
    def walk(self):
        print("walk")


class Dog(Mammal):
    def bark(self):
        print('bark')

class Cat(Mammal):
    def be ananying

dog1 = Dog()
dog1.bark()



import converters
from converters import kg_to_lbs

kg_to_lbs(80)

print(converters.kg_to_lbs(70))


def lbs_to_kg(weight):
    return weight * 0.45


def kg_to_lbs(weight):
    return weight / 0.45


def find_max(numbers):
    max = numbers[0]
    for number in numbers:
        if number > max:
            max = number
    return max


import utils
from utils import.find_max()


numbers = [10, 2, 6, 12]
max = find_max(numbers)
print(max)

numbers = [10, 2, 6, 12]
print(max)


from ecommerce import shipping

shipping.calc_shipping()


import random
for i in range(3):
    print(random.randint(80,100))

members = ['john','mary', 'bob','mosh']
leader = random.choice(members)
print(leader)

import random
class Dice:
    def roll(self):
        first = random.randint(1,6)
        second = random.randint(1,6)
        return first, second


dice =  Dice()
print(dice.roll())


from pathlib import Path

path = Path()
for file in path.glob('*'):
    print(file)



import openpyxl as xl
from openpyxl.chart import BarChart, Reference

def process_workbook(filename):

    wb = xl.load_workbook(filename)
    sheet = wb["Sheet1"]

    for row in range(2, sheet.max_row + 1):
        cell = sheet.cell(row, 3)
        correct_price = cell.value * 0.9
        correct_price_cell = sheet.cell(row, 4)
        correct_price_cell.value = correct_price

    value = Reference(sheet,
              min_row=2,
              max_row=sheet.max_row,
              min_col=4,
              max_col=4)

    chart = BarChart()
    chart.add_data(value)
    sheet.add_chart(chart, 'e2')


    wb.save(filename)

print()



i="zhangjinhui"
print(i)


print(3+15)

name = input('what is your name? ')
print('hello ' + name)



print(3+2)





name = 'ZHANG JIN HUI'
another_name = name.lower()
print(another_name)

name = input('what is your name? ')
print('Hello ' + name + '!')


name = input('what is your name? ')
print('hello there, ' + name)

numbers = [8,10,2,100000000]
print[:]


name = input('what is your name? ')
print('hey,there, '+name)



name = 'john'
is_new = False

print(name)


numbers = [2, 4, 6, 8, 6,  10]
number = numbers.sort()
number = numbers.reverse()
print(number)


number = [2, 4, 6, 8, 6,  10]
number1 = number.copy()
print(number1)


name = 'zhang jin hui'
name1 = name.upper()
print(name1)


email = '''
   hello,welcome to my team!
   
              enjoy it!!!

From Team Support!



'''
print(email)

for x in range(10):
for y in range(10):
    print(x,y)


prices = [0,30,125]

total = 0
for price in prices:
    total += price
    print(f"total: {total}")

    prices = [10, 20, 30]

    total = 0
    for price in prices:
        total += price
    print(f"total: {total}")



    prices = [124,123,215]
    total = 0
for price in prices:
    total += price
    print(f"total: {total}")


for x in range(1000):
for y in range(1000):
    print(x,y)

name = [2,3,5]
name1 = name.copy()
print(name1)


name = [1,2,5]
name1 = name.copy()
print(name1)

print('xxx')

xxx = 'xxxxx'
print(xxx)























